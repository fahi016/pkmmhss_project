"""
seating_logic.py
-----------------
Core logic for the Exam Seating Arrangement app.

Rule implemented:
    Within a single bench (a row of N seats/columns), two seats that are
    physically ADJACENT to each other must be occupied by students from
    DIFFERENT divisions (subjects/streams). Seats in the same column but
    a different bench row (i.e. front/back) are allowed to repeat the
    same division.

General approach:
    Columns of a bench are 2-coloured like a path graph:
        column index (1-based) -> bucket = (index - 1) % 2
    This guarantees adjacent columns always fall in different buckets.
    All divisions are then split between "Bucket 0" and "Bucket 1" such
    that the total number of students assigned to a bucket roughly
    matches the total seat-capacity available in that bucket across all
    rooms. Because the two buckets never share a division, adjacent
    seats are automatically guaranteed to differ.

    Within a bucket, students are handed out to rooms in order, and
    within a room they are placed round-robin across that room's
    bucket-columns, bench row by bench row -- this reproduces the same
    layout style used in the sample files provided by the school.
"""

from dataclasses import dataclass, field
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------- #
# Fixed division/stream mapping
# --------------------------------------------------------------------------- #
# This school's division scheme never changes:
#   S1, S2, S3, S4  -> Bio Science
#   S5, S8          -> Computer Science (S8 is +1 only, added 2026)
#   S6              -> Commerce
#   S7              -> Humanities
#
# For seating, Bio Science + Computer Science (S1-S5) always sit together on
# one side of the bench (the Left/Right-style columns), and Commerce +
# Humanities (S6, S7) always sit together on the other side (the
# Middle-style column) -- this matches the school's existing seating
# convention and is NOT re-balanced by student counts.

DIVISION_LABELS: Dict[str, str] = {
    "S1": "Bio Science",
    "S2": "Bio Science",
    "S3": "Bio Science",
    "S4": "Bio Science",
    "S5": "Computer Science",
    "S6": "Commerce",
    "S7": "Humanities",
    "S8": "Computer Science",
}

# Second-language codes in the SL column (separate from division/stream).
SECOND_LANGUAGE_LABELS: Dict[str, str] = {
    "A": "Arabic",
    "H": "Hindi",
    "M": "Malayalam",
    "U": "Urdu",
}

DIVISION_CODES = set(DIVISION_LABELS.keys())
SECOND_LANGUAGE_CODES = set(SECOND_LANGUAGE_LABELS.keys())

# Divisions that belong to the "Bio Science / Computer Science" seating group
# (Left & Right style columns in a 3-seat bench, or the bucket-0 columns more
# generally).
SCIENCE_GROUP_DIVISIONS = {"S1", "S2", "S3", "S4", "S5", "S8"}

# Divisions that belong to the "Commerce / Humanities" seating group
# (the Middle-style column in a 3-seat bench, or the bucket-1 columns more
# generally).
COMMERCE_HUMANITIES_GROUP_DIVISIONS = {"S6", "S7"}


def group_display_label(code: str) -> str:
    """Human-readable label for a seating group code (division or language)."""
    if code in DIVISION_LABELS:
        return DIVISION_LABELS[code]
    if code in SECOND_LANGUAGE_LABELS:
        return SECOND_LANGUAGE_LABELS[code]
    return code


def format_student_division(student: "Student") -> str:
    """
    Label for seating sheets and preview.
    Custom date-wise exam: 'Physics (S1)' — subject + division.
    Second Language exam: 'Arabic (S1)' — language name + class division.
    Regular Subjects: 'Bio Science (S1)' — stream name + division code.
    """
    class_div = student.class_division or student.division
    if student.subject:
        if class_div and class_div != student.subject:
            return f"{student.subject} ({class_div})"
        return student.subject
    if student.division in SECOND_LANGUAGE_LABELS:
        return f"{SECOND_LANGUAGE_LABELS[student.division]} ({class_div})"
    if class_div in DIVISION_LABELS:
        return f"{DIVISION_LABELS[class_div]} ({class_div})"
    return class_div


def roll_sort_key(roll_no: str):
    """Sort roll numbers numerically when possible."""
    try:
        return int(str(roll_no).strip())
    except (ValueError, TypeError):
        return str(roll_no)


def assign_roll_numbers_by_division(students: List["Student"],
                                    key_func=lambda s: s.class_division) -> None:
    """
    Assign roll numbers 1, 2, 3... within each division/class, following the
    order students appear in the input file. Counter resets when the division
    key changes (e.g. S1 students get 1-64, then S2 students restart at 1).
    """
    counters: Dict[str, int] = {}
    for s in students:
        key = key_func(s) or s.division or "UNASSIGNED"
        counters[key] = counters.get(key, 0) + 1
        s.roll_no = str(counters[key])


# --------------------------------------------------------------------------- #
# Data classes
# --------------------------------------------------------------------------- #

@dataclass
class Student:
    roll_no: str
    name: str
    division: str          # seating group for this exam (S1-S7, A/H/M/U, English, Subject, etc.)
    class_division: str = ""  # actual class/stream from DIVISION column (S1-S7)
    admn: str = ""
    subject: str = ""      # subject for custom date exam (e.g. Physics, Chemistry, etc.)
    extra: Dict[str, str] = field(default_factory=dict)

    def __repr__(self):
        return f"Student({self.name}, {self.division}, {self.roll_no})"


@dataclass
class RoomConfig:
    name: str
    num_benches: int
    seats_per_bench: int

    @property
    def capacity(self) -> int:
        return self.num_benches * self.seats_per_bench

    def bucket_columns(self):
        """Return (bucket0_columns, bucket1_columns) -- 1-based column indices."""
        b0, b1 = [], []
        for col in range(1, self.seats_per_bench + 1):
            if (col - 1) % 2 == 0:
                b0.append(col)
            else:
                b1.append(col)
        return b0, b1


@dataclass
class SeatAssignment:
    room: str
    bench_no: int
    column: int          # 1-based column index within the bench
    student: Optional[Student]


# --------------------------------------------------------------------------- #
# Seat-label helper
# --------------------------------------------------------------------------- #

def seat_label(column: int, seats_per_bench: int) -> str:
    """Human friendly label for a column position."""
    if seats_per_bench == 1:
        return "Only"
    if seats_per_bench == 2:
        return "Left" if column == 1 else "Right"
    if seats_per_bench == 3:
        return {1: "Left", 2: "Middle", 3: "Right"}[column]
    return f"Seat {column}"


# --------------------------------------------------------------------------- #
# Loading students from Excel
# --------------------------------------------------------------------------- #

def load_students_from_excel(path: str, sheet_name: str,
                              col_roll: str, col_name: str, col_division: str,
                              col_admn: Optional[str] = None,
                              col_class_division: Optional[str] = "DIVISION",
                              header_row: int = 1,
                              subject_mapping: Optional[Dict[str, str]] = None) -> List[Student]:
    """
    Read students from an Excel file.
    col_* parameters are the column HEADER TEXT (case-insensitive match)
    as found in the header_row of the given sheet.

    Roll numbers are NOT taken from col_roll. They are computed per class
    division (S1-S7): 1, 2, 3... in file order, restarting when the class
    division changes. col_class_division (default DIVISION) supplies the class.
    col_division supplies the seating group for the current exam session.
    If subject_mapping is provided (Custom mode), only divisions with a non-empty
    subject (not 'NO EXAM') are loaded, and student.subject is set.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    headers = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = cell.column

    def find_col(target, required=True):
        if target is None:
            return None
        key = target.strip().lower()
        if key not in headers:
            if required:
                raise ValueError(f"Column '{target}' not found in sheet '{sheet_name}'. "
                                  f"Available columns: {list(headers.keys())}")
            return None
        return headers[key]

    c_name = find_col(col_name)
    c_div = find_col(col_division)
    c_admn = find_col(col_admn, required=False) if col_admn else None
    c_class_div = find_col(col_class_division, required=False) if col_class_division else None

    students = []
    for row in ws.iter_rows(min_row=header_row + 1):
        name_val = row[c_name - 1].value
        if name_val is None or str(name_val).strip() == "":
            continue
        div_val = row[c_div - 1].value
        admn_val = row[c_admn - 1].value if c_admn else ""
        class_div_val = row[c_class_div - 1].value if c_class_div else div_val

        seating_division = str(div_val).strip() if div_val is not None else "UNASSIGNED"
        class_division = str(class_div_val).strip() if class_div_val is not None else seating_division

        subj = ""
        if subject_mapping is not None:
            mapped = subject_mapping.get(class_division, "")
            if not mapped or mapped.strip().upper() == "NO EXAM":
                continue
            subj = mapped.strip()
            seating_division = subj

        students.append(Student(
            roll_no="",
            name=str(name_val).strip(),
            division=seating_division,
            class_division=class_division,
            admn=str(admn_val).strip() if admn_val is not None else "",
            subject=subj,
        ))

    roll_key = (lambda s: s.class_division) if c_class_div else (lambda s: s.division)
    assign_roll_numbers_by_division(students, roll_key)
    return students


def get_sheet_names(path: str) -> List[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    return wb.sheetnames


def get_headers(path: str, sheet_name: str, header_row: int = 1) -> List[str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    return [str(c.value).strip() for c in ws[header_row] if c.value is not None]


# --------------------------------------------------------------------------- #
# Core assignment algorithm
# --------------------------------------------------------------------------- #

class SeatingResult:
    def __init__(self):
        self.assignments: List[SeatAssignment] = []   # every seat, incl. empty ones
        self.warnings: List[str] = []
        self.unseated: List[Student] = []

    def by_room(self) -> Dict[str, List[SeatAssignment]]:
        out: Dict[str, List[SeatAssignment]] = {}
        for a in self.assignments:
            out.setdefault(a.room, []).append(a)
        return out


def _partition_divisions(div_counts: Dict[str, int], target0: int, target1: int):
    """
    Greedily split divisions between bucket0 and bucket1 so their summed
    student counts approximate target0 / target1 respectively.
    Divisions are processed largest-first (classic LPT balancing).
    """
    items = sorted(div_counts.items(), key=lambda kv: -kv[1])
    bucket0, bucket1 = [], []
    sum0 = sum1 = 0
    for div, count in items:
        # Choose whichever bucket is further below its target (as a ratio),
        # preferring bucket0 on ties so results are deterministic.
        remaining0 = target0 - sum0
        remaining1 = target1 - sum1
        if remaining0 >= remaining1:
            bucket0.append(div)
            sum0 += count
        else:
            bucket1.append(div)
            sum1 += count
    return bucket0, bucket1


def division_sort_key(code: str):
    """Sort S1, S2, ... S10 numerically; other codes alphabetically."""
    c = str(code).strip().upper()
    if len(c) >= 2 and c[0] == "S" and c[1:].isdigit():
        return (0, int(c[1:]))
    return (1, c)


def parse_division_list(text: str) -> List[str]:
    """Parse 'S1, S2, S3' into ['S1', 'S2', 'S3'] sorted ascending."""
    codes = [p.strip().upper() for p in text.replace(";", ",").split(",") if p.strip()]
    return sorted(set(codes), key=division_sort_key)


def allocate_rooms_to_batches(rooms: List[RoomConfig],
                              batch_sizes: List[int]) -> List[List[RoomConfig]]:
    """
    Assign consecutive rooms to each batch in order.
    Non-final batches take rooms until capacity covers their student count.
    The final batch receives all remaining rooms.
    """
    if not batch_sizes:
        return []
    allocations: List[List[RoomConfig]] = []
    idx = 0
    for b, size in enumerate(batch_sizes):
        batch_rooms: List[RoomConfig] = []
        cap = 0
        if b == len(batch_sizes) - 1:
            batch_rooms = list(rooms[idx:])
            idx = len(rooms)
        else:
            while idx < len(rooms) and (cap < size or not batch_rooms):
                batch_rooms.append(rooms[idx])
                cap += rooms[idx].capacity
                idx += 1
        allocations.append(batch_rooms)
    return allocations


def _assign_seats_core(students: List[Student], rooms: List[RoomConfig],
                       sort_by_class_division: bool = False) -> SeatingResult:
    """Run the seating algorithm for one student set across the given rooms."""
    result = SeatingResult()

    by_div: Dict[str, List[Student]] = {}
    for s in students:
        by_div.setdefault(s.division, []).append(s)

    div_counts = {d: len(lst) for d, lst in by_div.items()}
    total_students = len(students)

    total_bucket0_cap = 0
    total_bucket1_cap = 0
    any_multi_seat_room = False
    for r in rooms:
        b0, b1 = r.bucket_columns()
        total_bucket0_cap += len(b0) * r.num_benches
        total_bucket1_cap += len(b1) * r.num_benches
        if r.seats_per_bench >= 2:
            any_multi_seat_room = True

    total_capacity = total_bucket0_cap + total_bucket1_cap

    if total_capacity < total_students:
        result.warnings.append(
            f"Total room capacity ({total_capacity}) is less than the number of "
            f"students ({total_students}). {total_students - total_capacity} "
            f"student(s) will be left unseated."
        )

    n_divisions = len(div_counts)
    if any_multi_seat_room and n_divisions < 2:
        safe_capacity = max(total_bucket0_cap, total_bucket1_cap)
        if total_students > safe_capacity:
            result.warnings.append(
                f"Only one seating group was detected in the data. To keep students "
                f"with the same paper apart, only {safe_capacity} of the {total_capacity} "
                f"seats can safely be used (one seat per bench is left empty as a spacer). "
                f"{total_students - safe_capacity} student(s) will not fit."
            )
        else:
            only_group = next(iter(div_counts.keys()), "?")
            only_label = group_display_label(only_group)
            group_desc = f"'{only_group}' ({only_label})" if only_label != only_group else f"'{only_group}'"
            result.warnings.append(
                f"Only one seating group was detected ({group_desc}), so one seat "
                f"per bench is intentionally left empty as a spacer. "
                f"Safe usable capacity: {safe_capacity} of {total_capacity} seats -- "
                f"this comfortably fits all {total_students} student(s). "
                f"For Regular Subjects use the DIVISION column (S1-S8). "
                f"For Second Language use the SL column (A/H/M/U). "
                f"For English turn ON 'Treat everyone as ONE single group'."
            )

    if total_capacity > 0:
        target0 = round(total_students * (total_bucket0_cap / total_capacity))
    else:
        target0 = 0
    target1 = total_students - target0

    divisions_present = set(div_counts.keys())
    known_stream_divisions = SCIENCE_GROUP_DIVISIONS | COMMERCE_HUMANITIES_GROUP_DIVISIONS
    if divisions_present and divisions_present.issubset(known_stream_divisions):
        bucket0_divs = sorted(d for d in divisions_present if d in SCIENCE_GROUP_DIVISIONS)
        bucket1_divs = sorted(d for d in divisions_present if d in COMMERCE_HUMANITIES_GROUP_DIVISIONS)
    else:
        bucket0_divs, bucket1_divs = _partition_divisions(div_counts, target0, target1)

    def build_queue(div_list):
        queue = []
        for d in sorted(div_list):
            queue.extend(by_div[d])
        if sort_by_class_division:
            queue.sort(key=lambda s: (division_sort_key(s.class_division or s.division),
                                        roll_sort_key(s.roll_no)))
        return queue

    queue0 = build_queue(bucket0_divs)
    queue1 = build_queue(bucket1_divs)

    idx0 = idx1 = 0

    for room in rooms:
        b0_cols, b1_cols = room.bucket_columns()
        for bench_no in range(1, room.num_benches + 1):
            for col in b0_cols:
                student = None
                if idx0 < len(queue0):
                    student = queue0[idx0]
                    idx0 += 1
                result.assignments.append(
                    SeatAssignment(room.name, bench_no, col, student)
                )
            for col in b1_cols:
                student = None
                if idx1 < len(queue1):
                    student = queue1[idx1]
                    idx1 += 1
                result.assignments.append(
                    SeatAssignment(room.name, bench_no, col, student)
                )

    result.unseated = queue0[idx0:] + queue1[idx1:]
    if result.unseated:
        names = ", ".join(s.name for s in result.unseated[:10])
        more = "" if len(result.unseated) <= 10 else f" (+{len(result.unseated)-10} more)"
        result.warnings.append(
            f"{len(result.unseated)} student(s) could not be seated due to "
            f"insufficient room capacity: {names}{more}"
        )

    return result


def assign_seats(students: List[Student], rooms: List[RoomConfig],
                 division_batches: Optional[List[List[str]]] = None) -> SeatingResult:
    """
    Assign students to seats.

    division_batches (Second Language only): each inner list is a batch of
    class divisions (e.g. [['S1','S2','S3','S5','S8'], ['S4','S6','S7']]). Batch 1
    is seated in the first rooms, batch 2 in the next, and so on. Divisions
    within a batch are filled in ascending order (S1, S2, ...).
    """
    if not division_batches:
        return _assign_seats_core(students, rooms)

    combined = SeatingResult()
    batch_sizes = []
    batch_students: List[List[Student]] = []

    all_assigned: set = set()
    for batch_divs in division_batches:
        div_set = set(batch_divs)
        batch = [s for s in students if (s.class_division or s.division) in div_set]
        batch.sort(key=lambda s: (division_sort_key(s.class_division or s.division),
                                  roll_sort_key(s.roll_no)))
        batch_students.append(batch)
        batch_sizes.append(len(batch))
        all_assigned.update(s.class_division or s.division for s in batch)

    present_divs = {s.class_division or s.division for s in students}
    missing = present_divs - all_assigned
    if missing:
        combined.warnings.append(
            f"Division batch setup does not include: {', '.join(sorted(missing, key=division_sort_key))}. "
            f"Those students will not be seated."
        )

    room_groups = allocate_rooms_to_batches(rooms, batch_sizes)

    for i, (batch, batch_rooms) in enumerate(zip(batch_students, room_groups), start=1):
        divs_label = ", ".join(sorted(division_batches[i - 1], key=division_sort_key))
        room_names = [r.name for r in batch_rooms]
        if not batch:
            combined.warnings.append(f"Batch {i} ({divs_label}): no students loaded.")
            continue
        if not batch_rooms:
            combined.warnings.append(
                f"Batch {i} ({divs_label}): no rooms allocated — "
                f"{len(batch)} student(s) could not be seated."
            )
            combined.unseated.extend(batch)
            continue

        batch_cap = sum(r.capacity for r in batch_rooms)
        combined.warnings.append(
            f"Batch {i} ({divs_label}): {len(batch)} student(s) -> "
            f"{', '.join(room_names)} ({batch_cap} seats)"
        )

        partial = _assign_seats_core(batch, batch_rooms, sort_by_class_division=True)
        combined.assignments.extend(partial.assignments)
        combined.warnings.extend(partial.warnings)
        combined.unseated.extend(partial.unseated)

    return combined


def verify_no_adjacent_same_division(result: SeatingResult, rooms: List[RoomConfig]) -> List[str]:
    """Sanity check: scans every bench row and confirms no two physically
    adjacent columns share the same division. Returns a list of violation
    descriptions (empty list = all good)."""
    violations = []
    room_map = {r.name: r for r in rooms}
    by_room = result.by_room()
    for room_name, seats in by_room.items():
        room = room_map[room_name]
        by_bench: Dict[int, Dict[int, SeatAssignment]] = {}
        for a in seats:
            by_bench.setdefault(a.bench_no, {})[a.column] = a
        for bench_no, cols in by_bench.items():
            for col in range(1, room.seats_per_bench):
                a1 = cols.get(col)
                a2 = cols.get(col + 1)
                if a1 and a2 and a1.student and a2.student:
                    if a1.student.division == a2.student.division:
                        violations.append(
                            f"{room_name} Bench {bench_no}: columns {col} & {col+1} "
                            f"both '{a1.student.division}' "
                            f"({a1.student.name} / {a2.student.name})"
                        )
    return violations


# --------------------------------------------------------------------------- #
# Excel export helpers
# --------------------------------------------------------------------------- #

TITLE_FONT = Font(bold=True, size=13)
SUBTITLE_FONT = Font(italic=True, size=10, color="555555")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
DIVISION_FILL = PatternFill(start_color="E8EEF3", end_color="E8EEF3", fill_type="solid")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header_row(ws, row_idx, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def export_seating_arrangement(result: SeatingResult, rooms: List[RoomConfig], path: str, exam_date: str = ""):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    by_room = result.by_room()

    # Detect whether this export uses the school's fixed Bio Science/Computer
    # Science vs Commerce/Humanities seating scheme, so the subtitle can
    # describe it explicitly (matching the school's existing sheet style).
    all_divs = sorted({a.student.division for a in result.assignments if a.student})
    known_stream_divisions = SCIENCE_GROUP_DIVISIONS | COMMERCE_HUMANITIES_GROUP_DIVISIONS
    fixed_scheme = bool(all_divs) and set(all_divs).issubset(known_stream_divisions)
    science_side = [d for d in all_divs if d in SCIENCE_GROUP_DIVISIONS]
    commerce_side = [d for d in all_divs if d in COMMERCE_HUMANITIES_GROUP_DIVISIONS]

    def _div_range(divs):
        if not divs:
            return ""
        return f"{divs[0]} to {divs[-1]}" if len(divs) > 1 else divs[0]

    def _stream_labels(divs):
        seen = []
        for d in divs:
            lbl = DIVISION_LABELS.get(d, d)
            if lbl not in seen:
                seen.append(lbl)
        return " / ".join(seen)

    for room in rooms:
        ws = wb.create_sheet(title=room.name[:31])
        header_title = f"EXAM SEATING ARRANGEMENT - {room.name.upper()}"
        if exam_date:
            header_title = f"EXAM SEATING ARRANGEMENT ({exam_date}) - {room.name.upper()}"
        ws.cell(row=1, column=1, value=header_title).font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=room.seats_per_bench + 1)

        if fixed_scheme and room.seats_per_bench >= 2:
            b0_cols, b1_cols = room.bucket_columns()
            b0_label = " & ".join(f"{seat_label(c, room.seats_per_bench)}({seat_label(c, room.seats_per_bench)[0]})" for c in b0_cols)
            b1_label = " & ".join(f"{seat_label(c, room.seats_per_bench)}({seat_label(c, room.seats_per_bench)[0]})" for c in b1_cols)
            subtitle = (
                f"Seating: {b0_label} - {_div_range(science_side)} ({_stream_labels(science_side)}) | "
                f"{b1_label} - {' & '.join(commerce_side)} ({_stream_labels(commerce_side)})"
            )
        else:
            cols_desc = ", ".join(
                f"{seat_label(c, room.seats_per_bench)}" for c in range(1, room.seats_per_bench + 1)
            )
            subtitle = f"Seats per bench: {cols_desc}"

        ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=room.seats_per_bench + 1)

        header_row = 4
        ws.cell(row=header_row, column=1, value="Bench No")
        for c in range(1, room.seats_per_bench + 1):
            ws.cell(row=header_row, column=c + 1, value=f"{seat_label(c, room.seats_per_bench)} Seat")
        _style_header_row(ws, header_row, room.seats_per_bench + 1)

        seats = by_room.get(room.name, [])
        by_bench: Dict[int, Dict[int, SeatAssignment]] = {}
        for a in seats:
            by_bench.setdefault(a.bench_no, {})[a.column] = a

        r = header_row + 1
        for bench_no in range(1, room.num_benches + 1):
            ws.cell(row=r, column=1, value=bench_no).alignment = CENTER
            ws.cell(row=r, column=1).border = BORDER
            for c in range(1, room.seats_per_bench + 1):
                a = by_bench.get(bench_no, {}).get(c)
                cell = ws.cell(row=r, column=c + 1)
                if a and a.student:
                    div = format_student_division(a.student)
                    cell.value = f"{a.student.name}\n({a.student.roll_no}) {div}"
                else:
                    cell.value = ""
                cell.alignment = CENTER
                cell.border = BORDER
            r += 1

        ws.column_dimensions["A"].width = 10
        for c in range(2, room.seats_per_bench + 2):
            ws.column_dimensions[get_column_letter(c)].width = 26
        for row in ws.iter_rows(min_row=header_row + 1, max_row=r - 1):
            ws.row_dimensions[row[0].row].height = 42

    # Summary sheet
    ws = wb.create_sheet(title="Summary", index=0)
    ws.cell(row=1, column=1, value="Exam Seating Arrangement - Summary").font = TITLE_FONT
    ws.cell(row=3, column=1, value="Room No")
    ws.cell(row=3, column=2, value="Total Students")
    ws.cell(row=3, column=3, value="Divisions Present")
    ws.cell(row=3, column=4, value="Details")
    _style_header_row(ws, 3, 4)

    r = 4
    grand_total = 0
    for room in rooms:
        seats = by_room.get(room.name, [])
        counts: Dict[str, int] = {}
        for a in seats:
            if a.student:
                counts[a.student.division] = counts.get(a.student.division, 0) + 1
        total = sum(counts.values())
        grand_total += total
        ws.cell(row=r, column=1, value=room.name)
        ws.cell(row=r, column=2, value=total)
        ws.cell(row=r, column=3, value=", ".join(sorted(counts.keys())))
        ws.cell(row=r, column=4, value=", ".join(f"{d}:{c}" for d, c in sorted(counts.items())))
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", wrap_text=True)
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=2, value=grand_total).font = Font(bold=True)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 40

    wb.save(path)


def _write_notice_sheet(wb, sheet_name: str, title: str,
                        assignments: List[SeatAssignment],
                        sort_key=None,
                        exam_date: str = ""):
    """Write one student-notice sheet and return the worksheet."""
    safe_name = sheet_name[:31] if sheet_name else "UNASSIGNED"
    ws = wb.create_sheet(title=safe_name)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    sub_title = "Please check your Room Number, Bench Number and Seat below"
    if exam_date:
        sub_title = f"Date: {exam_date}    |    {sub_title}"
    ws.cell(row=2, column=1, value=sub_title).font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = ["Sl No", "Roll No", "Student Name", "Room No", "Bench No", "Seat"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    _style_header_row(ws, 4, len(headers))

    if sort_key is None:
        sort_key = lambda a: roll_sort_key(a.student.roll_no)
    rows = sorted(assignments, key=sort_key)
    r = 5
    for sl, a in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=sl)
        ws.cell(row=r, column=2, value=a.student.roll_no)
        ws.cell(row=r, column=3, value=a.student.name)
        ws.cell(row=r, column=4, value=a.room)
        ws.cell(row=r, column=5, value=a.bench_no)
        ws.cell(row=r, column=6, value=a._label if hasattr(a, "_label") else "")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    return ws


def _notice_title_for_group(code: str, label: str) -> str:
    if code in SECOND_LANGUAGE_LABELS:
        return f"EXAM SEATING ARRANGEMENT NOTICE - SECOND LANGUAGE: {label} ({code})"
    if code in DIVISION_LABELS:
        return f"EXAM SEATING ARRANGEMENT NOTICE - DIVISION {code} ({label})"
    return f"EXAM SEATING ARRANGEMENT NOTICE - {code}" + (
        f" ({label})" if label and label != code else "")


def export_student_notice(result: SeatingResult, path: str,
                          division_labels: Optional[Dict[str, str]] = None,
                          include_class_division_notices: bool = False,
                          exam_date: str = ""):
    """
    Write student notice sheets to one workbook.

    By default one sheet per seating group (division for Regular Subjects,
    language code for Second Language, Subject for Custom). When
    include_class_division_notices is True, also adds S1-S7 division sheets.
    """
    division_labels = division_labels if division_labels else dict(DIVISION_LABELS)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    seated = [a for a in result.assignments if a.student]
    by_seating: Dict[str, List[SeatAssignment]] = {}
    for a in seated:
        by_seating.setdefault(a.student.division, []).append(a)

    by_language_exam = any(k in SECOND_LANGUAGE_LABELS for k in by_seating)

    for code in sorted(by_seating.keys()):
        label = division_labels.get(code, group_display_label(code))
        title = _notice_title_for_group(code, label)
        if exam_date and "NOTICE" in title:
            title = title.replace("NOTICE", f"NOTICE ({exam_date})")
        if by_language_exam and code in SECOND_LANGUAGE_LABELS:
            sort_key = lambda a: (
                a.student.class_division or "",
                roll_sort_key(a.student.roll_no),
            )
        else:
            sort_key = lambda a: roll_sort_key(a.student.roll_no)
        _write_notice_sheet(wb, code, title, by_seating[code], sort_key=sort_key, exam_date=exam_date)

    if include_class_division_notices:
        by_class: Dict[str, List[SeatAssignment]] = {}
        for a in seated:
            key = a.student.class_division or a.student.division
            by_class.setdefault(key, []).append(a)

        for code in sorted(by_class.keys()):
            label = DIVISION_LABELS.get(code, group_display_label(code))
            title = f"EXAM SEATING ARRANGEMENT NOTICE - DIVISION {code} ({label})"
            if exam_date:
                title = f"EXAM SEATING ARRANGEMENT NOTICE ({exam_date}) - DIVISION {code} ({label})"
            sheet_name = code if code not in by_seating else f"DIV-{code}"
            _write_notice_sheet(
                wb, sheet_name, title, by_class[code],
                sort_key=lambda a: roll_sort_key(a.student.roll_no),
                exam_date=exam_date,
            )

    wb.save(path)


def export_attendance_sheets(result: SeatingResult, rooms: List[RoomConfig], path: str,
                              date_columns: Optional[List[str]] = None,
                              group_labels: Optional[Dict[str, str]] = None,
                              exam_date: str = ""):
    if exam_date and not date_columns:
        date_columns = [exam_date]
    else:
        date_columns = date_columns or ["Date 1", "Date 2", "Date 3", "Date 4"]
    group_labels = group_labels or {}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    by_room = result.by_room()
    for room in rooms:
        ws = wb.create_sheet(title=room.name[:31])
        seats = [a for a in by_room.get(room.name, []) if a.student]
        sheet_title = f"EXAM ATTENDANCE SHEET ({exam_date})" if exam_date else "EXAM ATTENDANCE SHEET"
        ws.cell(row=1, column=1, value=sheet_title).font = TITLE_FONT
        ws.cell(row=2, column=1, value=f"{room.name.upper()}    |    Total Students: {len(seats)}").font = SUBTITLE_FONT
        ws.cell(row=3, column=1, value="Invigilator: ____________________").font = SUBTITLE_FONT

        base_headers = ["S.No", "Roll No", "Student Name", "Division", "Bench No"]
        headers = base_headers + date_columns
        header_row = 5
        for i, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=i, value=h)
        _style_header_row(ws, header_row, len(headers))

        by_div: Dict[str, List[SeatAssignment]] = {}
        for a in seats:
            by_div.setdefault(a.student.division, []).append(a)

        r = header_row + 1
        sno = 1
        for div in sorted(by_div.keys()):
            label = group_labels.get(div, group_display_label(div))
            header = f"GROUP: {div} ({label})" if label != div else f"GROUP: {div}"
            ws.cell(row=r, column=1, value=header).font = Font(bold=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            ws.cell(row=r, column=1).fill = DIVISION_FILL
            r += 1
            for a in sorted(by_div[div], key=lambda x: roll_sort_key(x.student.roll_no)):
                ws.cell(row=r, column=1, value=sno)
                ws.cell(row=r, column=2, value=a.student.roll_no)
                ws.cell(row=r, column=3, value=a.student.name)
                ws.cell(row=r, column=4, value=a.student.class_division or a.student.division)
                ws.cell(row=r, column=5, value=a.bench_no)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).border = BORDER
                r += 1
                sno += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        for i in range(len(base_headers) + 1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 14

    wb.save(path)


def export_question_paper_count(result: SeatingResult, rooms: List[RoomConfig], path: str, exam_date: str = ""):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Question Paper Count"

    all_divisions = sorted({a.student.division for a in result.assignments if a.student})

    title_text = "CLASS-WISE / SUBJECT-WISE STUDENT COUNT PER EXAM ROOM (For Question Paper Count)"
    if exam_date:
        title_text += f" - {exam_date}"
    ws.cell(row=1, column=1, value=title_text).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_divisions) + 2)

    header_row = 3
    ws.cell(row=header_row, column=1, value="Room No")
    for i, div in enumerate(all_divisions, start=2):
        ws.cell(row=header_row, column=i, value=div)
    ws.cell(row=header_row, column=len(all_divisions) + 2, value="Total")
    _style_header_row(ws, header_row, len(all_divisions) + 2)

    by_room = result.by_room()
    r = header_row + 1
    col_totals = {d: 0 for d in all_divisions}
    grand_total = 0
    for room in rooms:
        seats = [a for a in by_room.get(room.name, []) if a.student]
        counts: Dict[str, int] = {}
        for a in seats:
            counts[a.student.division] = counts.get(a.student.division, 0) + 1
        ws.cell(row=r, column=1, value=room.name)
        row_total = 0
        for i, div in enumerate(all_divisions, start=2):
            v = counts.get(div)
            if v:
                ws.cell(row=r, column=i, value=v)
                col_totals[div] += v
                row_total += v
        ws.cell(row=r, column=len(all_divisions) + 2, value=row_total)
        grand_total += row_total
        for c in range(1, len(all_divisions) + 3):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = CENTER
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for i, div in enumerate(all_divisions, start=2):
        ws.cell(row=r, column=i, value=col_totals[div]).font = Font(bold=True)
    ws.cell(row=r, column=len(all_divisions) + 2, value=grand_total).font = Font(bold=True)

    ws.column_dimensions["A"].width = 12
    for i in range(2, len(all_divisions) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 10

    wb.save(path)


def attach_seat_labels(result: SeatingResult, rooms: List[RoomConfig]):
    """Adds a human-readable seat label (Left/Middle/Right/Seat N) onto each
    SeatAssignment as `_label`, used by the notice export."""
    room_map = {r.name: r for r in rooms}
    for a in result.assignments:
        room = room_map[a.room]
        a._label = seat_label(a.column, room.seats_per_bench)
